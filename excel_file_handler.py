import openpyxl
import os
import time_handler
from openpyxl import load_workbook
from model.product import Product

WITHOUT_URL_MSG = 'SEM URL'
OUTPUT_DIR = "output"
START_ROW_PRODUCT = 3
START_COL_PRICE = 5
MAX_N_MARKETS = 5
HEADER_MKT_NAMES_CELLS = ['E1', 'F1', 'G1', 'H1', 'I1']


def read_products_from_file():
    wb = load_workbook('input.xlsx')
    active_workbook = wb['URLs Produtos']
    market_names = get_market_names(active_workbook)
    products_array = []
    i = START_ROW_PRODUCT
    while True:
        if active_workbook.cell(row=i, column=1).value is None:
            break
        product_attributes = []
        for j in range(1, 5):
            product_attributes.append(active_workbook.cell(row=i, column=j).value)
        product = Product(product_attributes[0], product_attributes[1], product_attributes[2], product_attributes[3])

        for j in range(5, 5+len(market_names)):
            product_url = active_workbook.cell(row=i, column=j).value
            product.add_market_info(market_names[j-5], product_url)

        products_array.append(product)
        i = i + 1
    return products_array


def get_market_names(active_workbook):
    market_names = []
    for col in range(5, 5 + MAX_N_MARKETS):
        market_name = active_workbook.cell(row=1, column=col).value
        if market_name is not None:
            market_names.append(market_name)
        else:
            break

    return market_names


def write_products_to_file(products_array):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=2, column=1).value = "Código"
    sheet.cell(row=2, column=2).value = "Item"
    sheet.cell(row=2, column=3).value = "Peso/Quantidade/Volume"
    sheet.cell(row=2, column=4).value = "Marca"

    total_price_list = []
    for p in range(0, MAX_N_MARKETS):
        total_price_list.append(0.0)

    market_names = products_array[0].get_market_names()
    max_price_offers = {}
    for market_name in market_names:
        max_price_offers[market_name] = 0

    i = START_ROW_PRODUCT
    for product in products_array:
        sheet.cell(row=i, column=1).value = product.code
        sheet.cell(row=i, column=2).value = product.name
        sheet.cell(row=i, column=3).value = product.measure
        sheet.cell(row=i, column=4).value = product.brand
        j = START_COL_PRICE
        for market_info in product.markets_info:
            if market_info['product_url'] is None:
                sheet.cell(row=i, column=j).value = WITHOUT_URL_MSG
            else:
                offers_qnt = len(market_info['offers'][0])
                market_name = market_info['market_name']
                if max_price_offers[market_name] < offers_qnt:
                    max_price_offers[market_name] = offers_qnt
                for z in range(0, offers_qnt):
                    price_info = market_info['offers'][0][z]
                    min_quantity = price_info['min_quantity']
                    price_float = float(price_info['price'])
                    price_str = '{:.2f}'.format(price_float)
                    if min_quantity > 1:
                        price_str = price_str + f' (x{min_quantity})'
                    sheet.cell(row=i, column=j).value = price_str
                    index = j - START_COL_PRICE
                    total_price_list[index] = total_price_list[index] + price_float
                    j = j + 1
        i = i + 1

    j = START_COL_PRICE
    next_ind_cell = 0
    for market_name in market_names:
        qnt_prices = max_price_offers[market_name]
        if qnt_prices > 1:
            shift_columns = qnt_prices - 1
            start_cell = HEADER_MKT_NAMES_CELLS[next_ind_cell]
            end_cell = HEADER_MKT_NAMES_CELLS[next_ind_cell + shift_columns]
            column_range = f'{start_cell}:{end_cell}'
            sheet.merge_cells(column_range)
            sheet.cell(row=1, column=j).value = str.upper(market_name)
            next_ind_cell = next_ind_cell + qnt_prices
            j = j + qnt_prices
        else:
            sheet.cell(row=1, column=j).value = str.upper(market_name)
            next_ind_cell = next_ind_cell + 1
            j = j + 1

    for col in range(START_COL_PRICE, j):
        sheet.cell(row=2, column=col).value = "Preço Unitário (R$)"

    sheet.cell(row=i, column=START_COL_PRICE-1).value = 'TOTAL'
    col_index = START_COL_PRICE
    for p in total_price_list:
        if p == 0.0:
            break
        sheet.cell(row=i, column=col_index).value = p
        col_index = col_index + 1

    current_path = os.getcwd()
    output_path = os.path.join(current_path, OUTPUT_DIR)
    if not os.path.isdir(output_path):
        os.mkdir(output_path)
    file_name = "OUTPUT_" + time_handler.get_now_time() + ".xlsx"
    output_file = os.path.join(output_path, file_name)
    wb.save(output_file)
    print()
    print(f'Excel file {output_file} was successfully generated!')


